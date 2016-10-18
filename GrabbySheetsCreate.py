# GrabbySheetCreate
# This program manipulates XLSX and XLS documents and creates worksheets based on an input file
# .1 --- GrabbySheetCreate created.
# ~~~~~~~~~~~~~~~~~~~~~~~~~ABSTRACT~~~~~~~~~~~~~~~~~~~~~~~~~
# This program opens an existing workbook and creates worksheets and names them based on an input file called DeviceInput.csv
#
import csv

# opens the csv file defined below.
#~~~~~~~~~~~~~~~~~~~CUSTOMIZATION AVAILABLE~~~~~~~~~~~~~~~~~~~
inputfile = open("DeviceInput.csv", 'rt')

# creates the reader object
reader = csv.reader(inputfile)
rownum = 0

# The following code imports Openpyxl libraries
from openpyxl import load_workbook

# The following code imports the existing workbook in order to manipulate the data within
#~~~~~~~~~~~~~~~~~~~CUSTOMIZATION AVAILABLE~~~~~~~~~~~~~~~~~~~
wb = load_workbook('config_template.xlsx')

# The following code defines ws as sheet 1 within the workbook
ws = wb.sheetnames[1]

# The following code prints the worksheet names in the workbook
#~~~~~~~~~~~~~~~~~~~CUSTOMIZATION AVAILABLE~~~~~~~~~~~~~~~~~~~

# Defines Headers in input file DeviceInput.csv
for row in reader:
    if rownum == 0:
        header = row
        hostname = header.index('hostname')
        active = header.index('Add')
    else:
        colnum = 0
        for col in row:
            if (colnum == header.index('Add')) and (col.lower() == 'y'):
                # ~~~~~~~~~~~~~~~~~~~CUSTOMIZATION AVAILABLE~~~~~~~~~~~~~~~~~~~
                # This is actually a list.
                devicedata = row
                # The following code defines the current worksheet for manipulation
                ws = wb.active

                # The following code defines the source worksheet we wish to copy
                # ~~~~~~~~~~~~~~~~~~~CUSTOMIZATION AVAILABLE~~~~~~~~~~~~~~~~~~~
                ws_copy = wb.copy_worksheet(wb['GatewayTemplate'])

                # The following code defines the new name of the newly copied worksheet
                ## This is an indexed variable in the list devicedata.  The brackets indicate which header is being referenced.
                # ~~~~~~~~~~~~~~~~~~~CUSTOMIZATION AVAILABLE~~~~~~~~~~~~~~~~~~~
                ws_copy.title = devicedata[hostname]

            colnum += 1 # Loop Counter
    rownum += 1 # Loop Counter

# The following code saves the worksheet
wb.save("config_template.xlsx")

print("These are the sheets in the following workbook {}".format(wb.sheetnames))





