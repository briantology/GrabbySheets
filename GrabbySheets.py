# GrabbySheet
# This program manipulates XLSX and XLS documents

# The following code imports Openpyxl libraries
from openpyxl import load_workbook

# The following code imports the existing workbook in order to manipulate the data within
wb = load_workbook('config_template.xlsx')

# The following code defines ws as sheet 1 within the workbook
ws = wb.sheetnames[1]

# The following code prints the worksheet names in the workbook
print("These are the sheets in the following workbook {}".format(wb.sheetnames))

# The following code defines the current worksheet for manipulation
ws = wb.active

# The following code defines the source worksheet we wish to copy
ws_copy = wb.copy_worksheet(wb['GatewayTemplate'])

# The following code defines the new name of the newly copied worksheet
## This should be a variable, but here it is hardcoded
ws_copy.title = 'NewTemplate'

# The following code creates new worksheets
# ws1 = wb.create_sheet(title="Pi2")
# ws2 = wb.create_sheet(title="Data")

# The following code saves the worksheet
# wb.save("config_template.xlsx")
