# GrabbySheet
# This program manipulates XLSX and XLS documents

# The following code imports Openpyxl libraries
from openpyxl import load_workbook

# The following code imports the existing workbook in order to manipulate the data within
wb = load_workbook('config_template.xlsx')

# The following code puts the worksheet names into a list of variables
sheets = wb.sheetnames


ws = sheets[2]

print(ws)

# The following code creates new worksheets
# ws1 = wb.create_sheet(title="Pi2")
# ws2 = wb.create_sheet(title="Data")

# The following code renames an existing worksheet which has been defined above
# ws1.title = "RenamedToPi3"



# The following code saves the worksheet
# wb.save("config_template.xlsx")
